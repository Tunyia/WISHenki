#!/usr/bin/env python3
"""Импорт студентов и начислений из Excel (лист «2-ой семестр 20252026»).

Модель:
  - Activity.base_reward = 0, is_completed = True
  - ActivityAttendance.bonus_points = сумма подкатегорий мероприятия для студента
  - students.total_points ← «Накоплены в 2-ом семестре» (col 4)
  - students.available_points ← «Доступны для списания» (col 5)

Запуск (Docker, из корня репозитория):
  docker compose exec api python import_excel.py --dry-run
  docker compose exec api python import_excel.py --force

Локально (venv в backend/project_vuz):
  python import_excel.py --file ../../excel_table/....xlsx --force
"""

from __future__ import annotations

import argparse
import re
import time
from dataclasses import dataclass, field
from pathlib import Path

from openpyxl import load_workbook
from sqlalchemy import text
from sqlalchemy.orm import Session

from core.attendance import upsert_attendance
from core.database import Base, SessionLocal, engine
from core.migrate import ensure_schema_updates
from models.activity import Activity, ActivityAttendance, ActivityEnrollment
from models.rating import Item, Student, Transaction, User

DEFAULT_FILENAME = "2 команда Копия Лидер инженерных школ 2025_2026.xlsx"
DEFAULT_SHEET = "2-ой семестр 20252026"
DATA_START_ROW = 3
NAME_COL = 2
GROUP_COL = 3
EARNED_COL = 4
AVAILABLE_COL = 5
FIRST_EVENT_COL = 6

_REPO_ROOT = Path(__file__).resolve().parent.parent.parent
_DEFAULT_FILE = _REPO_ROOT / "excel_table" / DEFAULT_FILENAME
_DOCKER_DEFAULT_FILE = Path("/excel_table") / DEFAULT_FILENAME


@dataclass
class EventColumnGroup:
    title: str
    columns: list[int]
    categories: list[str] = field(default_factory=list)


@dataclass
class ParsedStudent:
    full_name: str
    study_group: str
    earned_semester: int
    available: int
    event_bonuses: dict[str, int]  # event title -> bonus sum


@dataclass
class ParseResult:
    sheet_title: str
    events: list[EventColumnGroup]
    students: list[ParsedStudent]
    skipped_duplicate_rows: int
    warnings: list[str]


def wait_for_db(timeout_s: int = 30) -> None:
    deadline = time.time() + timeout_s
    last_err: Exception | None = None
    while time.time() < deadline:
        try:
            with engine.connect() as conn:
                conn.execute(text("SELECT 1"))
            return
        except Exception as e:
            last_err = e
            time.sleep(1)
    raise RuntimeError(f"DB is not ready after {timeout_s}s") from last_err


def ensure_schema() -> None:
    import models.activity  # noqa: F401
    import models.rating  # noqa: F401

    Base.metadata.create_all(bind=engine)
    ensure_schema_updates()


def parse_num(value) -> float:
    if value is None or value == "":
        return 0.0
    if isinstance(value, bool):
        return float(int(value))
    if isinstance(value, (int, float)):
        return float(value)
    s = str(value).strip().replace(",", ".")
    if not s or s in {"-", "—", "–"}:
        return 0.0
    try:
        return float(s)
    except ValueError:
        return 0.0


def normalize_title(value) -> str:
    return " ".join(str(value).split())


def extract_event_date(title: str) -> str:
    """Дата из заголовка мероприятия (если есть)."""
    m = re.search(
        r"(\d{1,2}[./]\d{1,2}(?:[./]\d{2,4})?|\d{1,2}\s*[-–]\s*\d{1,2}[./]\d{1,2}|"
        r"февраль|март|апрель|январь|Февраль-март)",
        title,
        re.IGNORECASE,
    )
    if m:
        return m.group(0).replace("  ", " ")
    return "2 семестр 2025/2026"


def infer_event_tags(categories: list[str]) -> list[str]:
    tags: list[str] = []
    mapping = {
        "организа": "Организация",
        "творческ": "Творческая деятельность",
        "медиа": "Медиа",
        "академ": "Академическая деятельность",
        "волонт": "Волонтерство",
        "професс": "Профессиональная деятельность",
    }
    for cat in categories:
        low = cat.lower()
        for key, tag in mapping.items():
            if key in low and tag not in tags:
                tags.append(tag)
    if not tags:
        tags.append("Социальное")
    return tags


def build_column_event_map(ws) -> dict[int, str]:
    col_event: dict[int, str] = {}
    for merge in ws.merged_cells.ranges:
        if merge.min_row == 1 and merge.max_row == 1:
            title = ws.cell(1, merge.min_col).value
            if title:
                t = normalize_title(title)
                for col in range(merge.min_col, merge.max_col + 1):
                    col_event[col] = t
    for col in range(1, ws.max_column + 1):
        if col not in col_event:
            value = ws.cell(1, col).value
            if value:
                col_event[col] = normalize_title(value)
    return col_event


def parse_event_groups(ws) -> list[EventColumnGroup]:
    col_event = build_column_event_map(ws)
    groups: list[EventColumnGroup] = []
    col = FIRST_EVENT_COL
    while col <= ws.max_column:
        if col not in col_event:
            col += 1
            continue
        title = col_event[col]
        start = col
        while col <= ws.max_column and col_event.get(col) == title:
            col += 1
        columns = list(range(start, col))
        categories = []
        for c in columns:
            raw = ws.cell(2, c).value
            cat = normalize_title(raw) if raw else f"Категория {c}"
            categories.append(cat)
        groups.append(EventColumnGroup(title=title, columns=columns, categories=categories))
    return groups


def parse_workbook(path: Path, sheet_name: str | None) -> ParseResult:
    wb = load_workbook(path, data_only=True, read_only=False)
    ws = None
    if sheet_name:
        if sheet_name not in wb.sheetnames:
            raise ValueError(
                f"Лист {sheet_name!r} не найден. Доступны: {', '.join(wb.sheetnames)}"
            )
        ws = wb[sheet_name]
    else:
        for candidate in wb.sheetnames:
            if "2" in candidate and "семестр" in candidate.lower():
                ws = wb[candidate]
                sheet_name = candidate
                break
        if ws is None:
            ws = wb.worksheets[1]
            sheet_name = ws.title

    events = parse_event_groups(ws)
    if not events:
        raise ValueError("Не найдены столбцы мероприятий (ожидались колонки с 6-й)")

    seen_keys: set[tuple[str, str]] = set()
    students: list[ParsedStudent] = []
    skipped = 0
    warnings: list[str] = []

    for row in range(DATA_START_ROW, ws.max_row + 1):
        raw_name = ws.cell(row, NAME_COL).value
        if raw_name is None or not str(raw_name).strip():
            continue
        full_name = normalize_title(raw_name)
        raw_group = ws.cell(row, GROUP_COL).value
        study_group = normalize_title(raw_group) if raw_group else ""
        key = (full_name, study_group)
        if key in seen_keys:
            skipped += 1
            continue
        seen_keys.add(key)

        earned = int(parse_num(ws.cell(row, EARNED_COL).value))
        available = int(parse_num(ws.cell(row, AVAILABLE_COL).value))
        event_bonuses: dict[str, int] = {}
        computed = 0

        for event in events:
            bonus = int(
                sum(parse_num(ws.cell(row, col).value) for col in event.columns)
            )
            if bonus > 0:
                event_bonuses[event.title] = bonus
            computed += bonus

        if earned != computed:
            warnings.append(
                f"Строка {row} ({full_name}): накоплено={earned}, "
                f"сумма по мероприятиям={computed}"
            )

        students.append(
            ParsedStudent(
                full_name=full_name,
                study_group=study_group,
                earned_semester=earned,
                available=available,
                event_bonuses=event_bonuses,
            )
        )

    return ParseResult(
        sheet_title=sheet_name or ws.title,
        events=events,
        students=students,
        skipped_duplicate_rows=skipped,
        warnings=warnings,
    )


def clear_all(db: Session) -> None:
    db.query(Transaction).delete()
    db.query(ActivityAttendance).delete()
    db.query(ActivityEnrollment).delete()
    db.query(User).delete()
    db.query(Student).delete()
    db.query(Item).delete()
    db.query(Activity).delete()
    db.commit()


def import_parsed(db: Session, data: ParseResult, *, dry_run: bool) -> dict:
    stats = {
        "students": 0,
        "activities": 0,
        "attendances": 0,
        "warnings": len(data.warnings),
        "skipped_duplicate_rows": data.skipped_duplicate_rows,
    }

    if dry_run:
        attendance_count = sum(
            len(s.event_bonuses) for s in data.students
        )
        stats["students"] = len(data.students)
        stats["activities"] = len(data.events)
        stats["attendances"] = attendance_count
        return stats

    title_to_activity: dict[str, Activity] = {}
    for event in data.events:
        tags = infer_event_tags(event.categories)
        act = Activity(
            title=event.title,
            organizer="ДПИШ",
            description=(
                f"Импорт из Excel («{data.sheet_title}»). "
                f"Категории: {', '.join(event.categories)}."
            ),
            categories=tags,
            base_reward=0,
            event_date=extract_event_date(event.title),
            images=[],
            is_completed=True,
        )
        db.add(act)
        db.flush()
        title_to_activity[event.title] = act
        stats["activities"] += 1

    name_group_to_student: dict[tuple[str, str], Student] = {}
    for row in data.students:
        student = Student(
            full_name=row.full_name,
            study_group=row.study_group,
            total_points=row.earned_semester,
            available_points=row.available,
        )
        db.add(student)
        db.flush()
        name_group_to_student[(row.full_name, row.study_group)] = student
        stats["students"] += 1

        for event_title, bonus in row.event_bonuses.items():
            activity = title_to_activity[event_title]
            upsert_attendance(db, activity, student, bonus_points=bonus)
            stats["attendances"] += 1

        # Баланс как в Excel (col4/col5), а не только пересчёт по посещениям
        student.total_points = row.earned_semester
        student.available_points = row.available

    db.commit()
    return stats


def resolve_default_file(explicit: str | None) -> Path:
    if explicit:
        path = Path(explicit)
        if not path.is_file():
            raise FileNotFoundError(f"Файл не найден: {path}")
        return path
    if _DOCKER_DEFAULT_FILE.is_file():
        return _DOCKER_DEFAULT_FILE
    if _DEFAULT_FILE.is_file():
        return _DEFAULT_FILE
    raise FileNotFoundError(
        f"Укажите --file. По умолчанию ожидался {_DEFAULT_FILE} "
        f"или {_DOCKER_DEFAULT_FILE} (Docker volume)."
    )


def print_report(data: ParseResult, stats: dict, *, dry_run: bool) -> None:
    mode = "DRY-RUN" if dry_run else "IMPORT"
    print(f"=== {mode} ===")
    print(f"Лист: {data.sheet_title}")
    print(f"Мероприятий: {len(data.events)}")
    print(f"Студентов (уникальных): {len(data.students)}")
    print(f"Пропущено дублей строк: {data.skipped_duplicate_rows}")
    print(f"Посещений с бонусом: {stats['attendances']}")
    if data.warnings:
        print(f"\nПредупреждения ({len(data.warnings)}):")
        for w in data.warnings[:20]:
            print(f"  - {w}")
        if len(data.warnings) > 20:
            print(f"  ... и ещё {len(data.warnings) - 20}")
    if dry_run:
        print("\nПримеры мероприятий:")
        for ev in data.events[:5]:
            print(f"  • {ev.title[:70]} ({len(ev.columns)} кат.)")
        with_pts = [s for s in data.students if s.event_bonuses]
        print(f"\nСтудентов с начислениями за мероприятия: {len(with_pts)}")
        if with_pts:
            s = with_pts[0]
            print(
                f"  Пример: {s.full_name} — накоплено {s.earned_semester}, "
                f"доступно {s.available}, мероприятий {len(s.event_bonuses)}"
            )


def main() -> None:
    parser = argparse.ArgumentParser(
        description="Импорт студентов и бонусных вишенок из Excel"
    )
    parser.add_argument(
        "--file",
        default=None,
        help=f"Путь к .xlsx (по умолчанию excel_table/{DEFAULT_FILENAME})",
    )
    parser.add_argument(
        "--sheet",
        default=DEFAULT_SHEET,
        help=f"Имя листа (по умолчанию {DEFAULT_SHEET!r})",
    )
    parser.add_argument(
        "--force",
        action="store_true",
        help="Очистить students/activities/users и импортировать заново",
    )
    parser.add_argument(
        "--dry-run",
        action="store_true",
        help="Только разбор и отчёт, без записи в БД",
    )
    parser.add_argument("--wait", type=int, default=30, help="Секунд ждать БД")
    args = parser.parse_args()

    path = resolve_default_file(args.file)
    print(f"Файл: {path}")

    data = parse_workbook(path, args.sheet)

    if args.dry_run:
        stats = {
            "students": len(data.students),
            "activities": len(data.events),
            "attendances": sum(len(s.event_bonuses) for s in data.students),
            "warnings": len(data.warnings),
            "skipped_duplicate_rows": data.skipped_duplicate_rows,
        }
        print_report(data, stats, dry_run=True)
        return

    wait_for_db(timeout_s=args.wait)
    ensure_schema()

    db = SessionLocal()
    try:
        if args.force:
            print("Очистка существующих данных...")
            clear_all(db)
        stats = import_parsed(db, data, dry_run=False)
        print_report(data, stats, dry_run=False)
        print("\nИмпорт завершён успешно.")
    finally:
        db.close()


if __name__ == "__main__":
    main()
